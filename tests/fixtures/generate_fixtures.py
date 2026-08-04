"""前処理機能を網羅的に検証するためのテストデータ生成スクリプト。

実行方法:
    python tests/fixtures/generate_fixtures.py

再実行すれば毎回同じ内容のファイルを冪等に再生成する。
各関数は tests/README.md のテストファイル設計表に対応する。
"""

import csv
import os

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))


def _path(name: str) -> str:
    return os.path.join(HERE, name)


def _new_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _write_rows(ws, rows, start_row: int = 1, start_col: int = 1) -> None:
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            if v is not None:
                ws.cell(row=start_row + i, column=start_col + j, value=v)


def _write_csv(name: str, rows) -> None:
    with open(_path(name), "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow(["" if v is None else v for v in row])


# ---------------------------------------------------------------------------
# 1. step2_structure_test.xlsx
# ---------------------------------------------------------------------------
def gen_step2_structure_test():
    wb = _new_wb()

    # --- シート1: ヘッダー種別（A1, B1, C1両分岐） ---
    ws = wb.create_sheet("ヘッダー種別")
    # (a) タイトル行 + 単一ヘッダー
    _write_rows(
        ws,
        [
            ["2024年度 支店別売上実績"],
            ["支店", "売上"],
            ["東京", 1000],
            ["大阪", 800],
        ],
        start_row=1,
    )
    # (b) 結合セルスパン型2段ヘッダー（横方向の結合セル）
    # 品質フィルタ（密なデータ行2行以上）を満たすためデータ行を2行にする
    _write_rows(
        ws,
        [
            ["商品", "東京", None, "大阪", None],
            [None, "売上", "原価", "売上", "原価"],
            ["家電", 100, 70, 80, 50],
            ["衣料", 90, 60, 70, 45],
        ],
        start_row=7,
    )
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
    ws.merge_cells(start_row=7, start_column=4, end_row=7, end_column=5)
    # (c) 名称+単位ペア2段ヘッダー
    _write_rows(
        ws,
        [
            ["支店", "従業者数"],
            [None, "人"],
            ["東京", 50],
            ["大阪", 30],
        ],
        start_row=13,
    )

    # --- シート2: 複数テーブル配置（B2, B3, B4） ---
    ws2 = wb.create_sheet("複数テーブル配置")
    # 横並び2テーブル（間に2列の空白）
    _write_rows(ws2, [["支店", "売上"], ["東京", 100], ["大阪", 80]], start_row=1, start_col=1)
    _write_rows(ws2, [["地域", "人口"], ["関東", 4000], ["関西", 2000]], start_row=1, start_col=5)
    # 縦方向: 2行空けてメモ + テーブル + 末尾注記
    _write_rows(
        ws2,
        [
            ["※本データは概算値です"],
            ["部門", "人数"],
            ["営業", 30],
            ["総務", 10],
            ["出典：自社調査"],
        ],
        start_row=6,
    )
    # 数値を一切含まない小さく幅広いテキストのみのブロック。
    # 品質フィルタは数値の有無だけでは discard しない（正当な短いテキスト表
    # として認識されるべき）ため、このブロックは正しく検出されることを
    # 期待する（B4改善の回帰テスト、tests/test_step2_detect.py参照）。
    _write_rows(
        ws2,
        [
            ["項目A", "項目B", "項目C", "項目D", "項目E"],
            ["a", "b", "c", "d", "e"],
            ["f", "g", "h", "i", "j"],
        ],
        start_row=15,
    )

    # 低品質領域（B4: 数値がほぼ無く長文セルが多い、非表形式のプロース
    # ブロック。_classify_table_quality の長文チェックで引き続き discard
    # されるべきケース）
    long_a = "これは表の実データではなく注記として扱われるべき長い説明文の例です。単なる参考情報として記載しています。"
    long_b = "こちらも同様の趣旨で記載された長めの補足説明文であり参考情報です。データの値としては扱いません。"
    _write_rows(
        ws2,
        [
            ["説明1", "説明2"],
            [long_a, long_b],
            [long_b, long_a],
        ],
        start_row=20,
    )

    # --- シート3: 新規対応レイアウト（1列ギャップ横並び／1行テーブル／
    #              幅広テーブル中の全文字列行混在） ---
    ws3 = wb.create_sheet("新規対応レイアウト")
    # 1列だけ空けた横並び2テーブル（gap_threshold=1で分離できることを確認）
    _write_rows(ws3, [["支店", "売上"], ["東京", 100], ["大阪", 80]], start_row=1, start_col=1)
    _write_rows(ws3, [["地域", "人口"], ["関東", 4000], ["関西", 2000]], start_row=1, start_col=4)
    # 1行だけのテーブル（2行空けて配置）
    _write_rows(ws3, [["商品", "在庫"], ["家電", 5]], start_row=6, start_col=1)
    # 幅広（5列）テーブルの中に全文字列の行が混在する（分断されず1テーブルの
    # ままであることを確認）
    _write_rows(
        ws3,
        [
            ["支店", "値1", "値2", "値3", "値4"],
            ["東京", 100, 90, 80, 70],
            ["大阪", 60, 50, 40, 30],
            ["名古屋", "テキスト1", "テキスト2", "テキスト3", "テキスト4"],
            ["福岡", 20, 10, 5, 1],
            ["札幌", 15, 12, 8, 3],
        ],
        start_row=10,
    )

    wb.save(_path("step2_structure_test.xlsx"))


# ---------------------------------------------------------------------------
# 2. step1_csv_parsing_test.csv
# ---------------------------------------------------------------------------
def gen_step1_csv_parsing_test():
    rows = [
        ["支店", "売上", "備考"],
        ["東京", "1000", " 順調"],
        ["大阪", "800", ""],
        [" 名古屋 ", "600.5", "微減"],
        ["", "", ""],
    ]
    _write_csv("step1_csv_parsing_test.csv", rows)


# ---------------------------------------------------------------------------
# 3. step1_excel_cell_indent_test.xlsx
# ---------------------------------------------------------------------------
def gen_step1_excel_cell_indent_test():
    wb = _new_wb()
    ws = wb.create_sheet("Sheet1")
    rows = [
        ("地域", "売上高", None),
        ("関東地方", 1000000, 0),
        ("東京都", 500000, 1),
        ("新宿区", 300000, 2),
        ("渋谷区", 200000, 2),
        ("関西地方", 500000, 0),
        ("大阪府", 500000, 1),
    ]
    for i, (label, val, indent) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)
        if indent is not None:
            ws.cell(row=i, column=1).alignment = Alignment(indent=indent)
    # 通常列（インデントなし、回帰確認用）
    for i in range(1, len(rows) + 1):
        ws.cell(row=i, column=3, value="通常" if i > 1 else "備考")
    wb.save(_path("step1_excel_cell_indent_test.xlsx"))


# ---------------------------------------------------------------------------
# 4. step3_pivot_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_pivot_test():
    wb = _new_wb()
    ws = wb.create_sheet("Pivot")
    rows = [
        ["支店", "属性", "値"],
        ["東京", "売上", 100],
        ["東京", "利益", 20],
        ["大阪", "売上", 80],
        ["大阪", "利益", 15],
    ]
    _write_rows(ws, rows)
    wb.save(_path("step3_pivot_test.xlsx"))


# ---------------------------------------------------------------------------
# 4b. step3_pivot_scale_test.xlsx（変換規模の事前予測機能: 大規模Pivot検出用）
# ---------------------------------------------------------------------------
def gen_step3_pivot_scale_test():
    # PIVOT_PERFORMANCE_WARNING_THRESHOLD（200）を超える201種類の属性を持つ
    # KVペア表。2つのキー（拠点）× 201属性 = 402行。normalize_tables()が
    # このテーブルのPivotを保留し、以降のStep3処理を完全にスキップすることを
    # end-to-endで検証するためのフィクスチャ。
    wb = _new_wb()
    ws = wb.create_sheet("PivotScale")
    n_attrs = 201
    rows = [["拠点", "属性", "値"]]
    for key in ["拠点A", "拠点B"]:
        for i in range(n_attrs):
            rows.append([key, f"属性{i:03d}", i])
    _write_rows(ws, rows)
    wb.save(_path("step3_pivot_scale_test.xlsx"))


# ---------------------------------------------------------------------------
# 5. step3_multiaxis_header_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_multiaxis_header_test():
    wb = _new_wb()
    ws = wb.create_sheet("MultiAxis")
    # ラベル列位置（1列目）は両ヘッダー行とも空白にする。
    # apply_multi_axis_header の label_positions は「全ての軸候補行で空白の列」
    # を基準にラベル列と判定するため、片方の行にでも値を置くとその列は
    # 軸データ位置として扱われてしまい、ラベル列として認識されない。
    rows = [
        [None, "売上", "売上", "原価", "原価"],
        [None, "東京", "大阪", "東京", "大阪"],
        ["2024年", 100, 80, 60, 45],
        ["2025年", 120, 90, 65, 50],
    ]
    _write_rows(ws, rows)
    wb.save(_path("step3_multiaxis_header_test.xlsx"))


# ---------------------------------------------------------------------------
# 6. step3_column_hierarchy_split_test.csv
# ---------------------------------------------------------------------------
def gen_step3_column_hierarchy_split_test():
    rows = [
        ["支店", "帳票プラン1 / 帳票プラン2", "T/O比"],
        ["東京支店", "フレッツ光 / おまかせLAN構築", 12.5],
        ["大阪営業所", "光コラボ / リモートアクセス", 9.7],
        ["名古屋支店", "フレッツ光 / セキュリティ対策", 15.1],
    ]
    _write_csv("step3_column_hierarchy_split_test.csv", rows)


# ---------------------------------------------------------------------------
# 7. step3_paren_annotation_test.csv
# ---------------------------------------------------------------------------
def gen_step3_paren_annotation_test():
    rows = [
        ["支店", "人口"],
        ["東京(本社)", "15歳以上人口(人)"],
        ["大阪(支店)", "労働力率(％)"],
        ["名古屋(支店)", "15歳以上人口(人)"],
        ["福岡(支店)", "労働力率(％)"],
    ]
    _write_csv("step3_paren_annotation_test.csv", rows)


# ---------------------------------------------------------------------------
# 8. step3_hierarchy_expand_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_hierarchy_expand_test():
    wb = _new_wb()

    # シート1: インデント方式（先頭半角スペース、literal）
    ws1 = wb.create_sheet("インデント方式")
    rows1 = [
        ["地域", "売上"],
        ["東日本", 1600],
        ["  神奈川事業部", 1100],
        ["    横浜支店", 700],
        ["    川崎支店", 400],
        ["  千葉事業部", 500],
    ]
    _write_rows(ws1, rows1)

    # シート2: 区切り文字方式
    ws2 = wb.create_sheet("区切り文字方式")
    rows2 = [
        ["地域", "売上"],
        ["東日本＞神奈川事業部＞横浜支店", 700],
        ["東日本＞静岡事業部", 400],
        ["東日本＞静岡事業部＞浜松支店", 250],
    ]
    _write_rows(ws2, rows2)

    # シート3: セル書式インデント方式（Excelの「インデントを増やす」）
    ws3 = wb.create_sheet("セル書式インデント方式")
    rows3 = [
        ("地域", "売上", None),
        ("東日本", 1100, 0),
        ("神奈川事業部", 1100, 1),
        ("横浜支店", 700, 2),
        ("川崎支店", 400, 2),
    ]
    for i, (label, val, indent) in enumerate(rows3, start=1):
        ws3.cell(row=i, column=1, value=label)
        ws3.cell(row=i, column=2, value=val)
        if indent is not None:
            ws3.cell(row=i, column=1).alignment = Alignment(indent=indent)

    wb.save(_path("step3_hierarchy_expand_test.xlsx"))


# ---------------------------------------------------------------------------
# 9. step3_fill_uchi_aggregate_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_fill_uchi_aggregate_test():
    wb = _new_wb()

    # シート1: 前方補完 + 集計行除去 + 階層整合性検証（PASS/SHORTFALL/EXCESS）
    # 「うち」書きを混ぜると、うち分離が内訳行を別テーブルへ分離してしまい
    # 集計行の検証対象（本体に残る個別行）がずれるため、うち分離とは別の
    # シートに分けて検証条件を明確にする。
    ws1 = wb.create_sheet("前方補完と集計除去")
    rows1 = [
        ["支店", "部門", "売上"],
        ["東京", "営業", 60],
        [None, "総務", 40],
        [None, "合計", 100],   # 東京: 60+40=100 と一致 → PASS
        ["大阪", "営業", 30],
        [None, "総務", 20],
        [None, "合計", 45],    # 大阪: 30+20=50 > 45 → EXCESS（二重計上または単位ミス）
        ["福岡", "営業", 50],
        [None, "総務", 60],
        [None, "合計", 150],   # 福岡: 50+60=110 < 150 → SHORTFALL（内訳あり）
    ]
    _write_rows(ws1, rows1)

    # シート2: 「うち」書き識別と別テーブル分離（単独で検証）
    ws2 = wb.create_sheet("うち分離")
    rows2 = [
        ["支店", "区分", "人数"],
        ["東京", "社員", 100],
        [None, "うち男性", 60],
        [None, "うち女性", 40],
        ["大阪", "社員", 50],
        [None, "うち男性", 20],
        [None, "うち女性", 20],
    ]
    _write_rows(ws2, rows2)

    wb.save(_path("step3_fill_uchi_aggregate_test.xlsx"))


# ---------------------------------------------------------------------------
# 10. step3_unit_split_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_unit_split_test():
    wb = _new_wb()
    ws = wb.create_sheet("Sheet1")
    rows = [
        ["年", "指標", "値"],
        ["2024", "人口(人)", 1200],
        ["2024", "労働力率(％)", 62.5],
        ["2025", "人口(人)", 1250],
        ["2025", "労働力率(％)", 63.0],
    ]
    _write_rows(ws, rows)
    wb.save(_path("step3_unit_split_test.xlsx"))


# ---------------------------------------------------------------------------
# 11. step3_invalid_columns_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_invalid_columns_test():
    wb = _new_wb()
    ws = wb.create_sheet("Sheet1")
    # 値=全欠損列、Unnamed: 3=無名だがデータあり、支店(2列目)=重複カラム名。
    # 各データ行の数値比率が0.30を超えるよう支店コード（数値）・件数（数値）
    # を混ぜ、Step2の行分類でヘッダー行と誤判定されない（品質フィルタで
    # discardされない）ようにする。
    #
    # 「無名＋全欠損」列（列名もデータも一切無い列）はここには含めない。
    # そのような列は raw grid 上で複数テーブルの1列ギャップ区切り
    # （gap_threshold=1）と本質的に区別が付かず、この列単体でテーブルが
    # 分割される可能性があるため、そのケースは
    # test_step3_normalize_determ.py::test_detect_invalid_columns で
    # detect_invalid_columns() に直接 DataFrame を渡して個別に検証する。
    rows = [
        ["支店", "支店", "値", "Unnamed: 3", "件数"],
        ["東京", 1001, None, "参考A", 100],
        ["大阪", 1002, None, None, 200],
    ]
    _write_rows(ws, rows)
    wb.save(_path("step3_invalid_columns_test.xlsx"))


# ---------------------------------------------------------------------------
# 12. step3_wide_to_long_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_wide_to_long_test():
    wb = _new_wb()

    # シート1: Tier1 時系列語彙
    ws1 = wb.create_sheet("時系列語彙")
    rows1 = [
        ["支店", "2023_売上", "2023_原価", "2024_売上", "2024_原価"],
        ["東京", 100, 60, 120, 65],
        ["大阪", 80, 50, 90, 55],
    ]
    _write_rows(ws1, rows1)

    # シート2: Tier2 区切り文字
    # Tier2は区切り文字による分解のため語彙の裏付けがなく、閾値が厳しい
    # （一致率0.9以上）。非分解列（商品）が1列混ざるだけで比率が下がるため、
    # 支社を5つに増やして10/11≒0.91で閾値を満たす。
    ws2 = wb.create_sheet("区切り文字")
    branches = ["東京支社", "大阪支社", "名古屋支社", "福岡支社", "札幌支社"]
    header2 = ["商品"] + [f"{b}_{ind}" for b in branches for ind in ("売上", "原価")]
    data2a = ["家電"] + [v for b in branches for v in (100, 70)]
    data2b = ["衣料"] + [v for b in branches for v in (90, 60)]
    rows2 = [header2, data2a, data2b]
    _write_rows(ws2, rows2)

    # シート3: Tier3 区切りなし複合語（LLM要）
    ws3 = wb.create_sheet("LLM要")
    rows3 = [
        ["月", "ゴールド利用額", "ゴールド来店数", "シルバー利用額", "シルバー来店数", "ブロンズ利用額", "ブロンズ来店数"],
        ["4月", 500, 12, 300, 8, 150, 5],
        ["5月", 520, 13, 310, 9, 160, 6],
    ]
    _write_rows(ws3, rows3)

    wb.save(_path("step3_wide_to_long_test.xlsx"))


# ---------------------------------------------------------------------------
# 13. step3_crosstab_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_crosstab_test():
    wb = _new_wb()
    ws = wb.create_sheet("Sheet1")
    rows = [
        ["支店", "2023年", "2024年", "2025年"],
        ["東京", 100, 120, 130],
        ["大阪", 80, 90, 95],
    ]
    _write_rows(ws, rows)
    wb.save(_path("step3_crosstab_test.xlsx"))


# ---------------------------------------------------------------------------
# 14. 2025年度_サービスX_実績.xlsx
# ---------------------------------------------------------------------------
def gen_external_metadata_test():
    wb = _new_wb()
    ws = wb.create_sheet("支店別（契約数）")
    rows = [
        ["区分1", "4月", "5月"],
        ["ライトプラン", 20, 22],
        ["スタンダードプラン", 35, 38],
    ]
    _write_rows(ws, rows)
    wb.save(_path("2025年度_サービスX_実績.xlsx"))


# ---------------------------------------------------------------------------
# 15. step3_transpose_test.xlsx
# ---------------------------------------------------------------------------
def gen_step3_transpose_test():
    wb = _new_wb()
    ws = wb.create_sheet("Sheet1")
    rows = [
        ["指標", "東京", "大阪"],
        ["売上", 100, 80],
        ["利益", 20, 15],
    ]
    _write_rows(ws, rows)
    wb.save(_path("step3_transpose_test.xlsx"))


# ---------------------------------------------------------------------------
# 16. step4_sum_relations_test.xlsx
# ---------------------------------------------------------------------------
def gen_step4_sum_relations_test():
    wb = _new_wb()
    cols = ["区分", "4月", "5月"]
    # 親: 全社計 = 東京+大阪+名古屋（要素ごとに一致）。品質フィルタが要求する
    # 「密なデータ行2行以上」を満たすため、各シートに2行ずつ用意する。
    ws_p = wb.create_sheet("全社計")
    _write_rows(ws_p, [cols, ["A", 300, 330], ["B", 150, 165]])
    ws_c1 = wb.create_sheet("東京")
    _write_rows(ws_c1, [cols, ["A", 100, 110], ["B", 50, 55]])
    ws_c2 = wb.create_sheet("大阪")
    _write_rows(ws_c2, [cols, ["A", 120, 130], ["B", 60, 66]])
    ws_c3 = wb.create_sheet("名古屋")
    _write_rows(ws_c3, [cols, ["A", 80, 90], ["B", 40, 44]])
    wb.save(_path("step4_sum_relations_test.xlsx"))


# ---------------------------------------------------------------------------
# 17. step4_sheet_levels_test.xlsx
# ---------------------------------------------------------------------------
def gen_step4_sheet_levels_test():
    wb = _new_wb()
    cols = ["支店", "帳票集計区分1", "4月", "5月"]

    # 系統1: スケールが大きいだけの明細シート群（誤検出を防ぐべきケース）
    def _detail_rows(base):
        return [cols] + [
            ["東京", "A", base * 1, base * 1],
            ["大阪", "A", base * 2, base * 2],
            ["名古屋", "A", base * 3, base * 3],
        ]

    ws_ap = wb.create_sheet("AP数")
    _write_rows(ws_ap, _detail_rows(10))
    ws_plan = wb.create_sheet("プラン数")
    _write_rows(ws_plan, _detail_rows(5))
    ws_id = wb.create_sheet("ID数")
    _write_rows(ws_id, _detail_rows(100))  # 値が大きいだけで他シートとの合計関係はない

    # 系統2: 真の集計関係（全社計 = 東京+大阪+名古屋、系統1とは別スキーマ）。
    # 品質フィルタが要求する「密なデータ行2行以上」を満たすため2指標分用意する。
    cols2 = ["部門", "指標", "値"]
    ws_total = wb.create_sheet("全社計2")
    _write_rows(ws_total, [cols2, ["全体", "件数", 600], ["全体", "金額", 6000]])
    ws_t1 = wb.create_sheet("東京2")
    _write_rows(ws_t1, [cols2, ["東京", "件数", 100], ["東京", "金額", 1000]])
    ws_t2 = wb.create_sheet("大阪2")
    _write_rows(ws_t2, [cols2, ["大阪", "件数", 200], ["大阪", "金額", 2000]])
    ws_t3 = wb.create_sheet("名古屋2")
    _write_rows(ws_t3, [cols2, ["名古屋", "件数", 300], ["名古屋", "金額", 3000]])

    wb.save(_path("step4_sheet_levels_test.xlsx"))


def main():
    generators = [
        gen_step2_structure_test,
        gen_step1_csv_parsing_test,
        gen_step1_excel_cell_indent_test,
        gen_step3_pivot_test,
        gen_step3_pivot_scale_test,
        gen_step3_multiaxis_header_test,
        gen_step3_column_hierarchy_split_test,
        gen_step3_paren_annotation_test,
        gen_step3_hierarchy_expand_test,
        gen_step3_fill_uchi_aggregate_test,
        gen_step3_unit_split_test,
        gen_step3_invalid_columns_test,
        gen_step3_wide_to_long_test,
        gen_step3_crosstab_test,
        gen_external_metadata_test,
        gen_step3_transpose_test,
        gen_step4_sum_relations_test,
        gen_step4_sheet_levels_test,
    ]
    for gen in generators:
        gen()
        print(f"generated: {gen.__name__}")
    print(f"\n{len(generators)} files generated in {HERE}")


if __name__ == "__main__":
    main()
