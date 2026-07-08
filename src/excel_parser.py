import io
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd

from .models import DetectedTable
from .table_formatter import (
    detect_cross_table, detect_header_roles, fill_grouping_cols,
    merge_header_rows, remove_aggregates, stack_cross_table,
)


# ---------------------------------------------------------------------------
# Grid 構築
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# 数式評価ヘルパー
# ---------------------------------------------------------------------------


def _col_to_num(col: str) -> int:
    """'A'→1, 'Z'→26, 'AA'→27（大文字・小文字を区別しない）。"""
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _parse_cell_ref(ref: str) -> Tuple[Optional[int], Optional[int]]:
    """'D5', '$D$5' → (row=5, col=4)。失敗時は (None, None) を返す。"""
    m = re.match(r"\$?([A-Z]+)\$?(\d+)$", ref.strip().upper())
    if m:
        return int(m.group(2)), _col_to_num(m.group(1))
    return None, None


def _eval_formula(formula: str, grid: List[List[Any]]) -> Optional[float]:
    """
    構築済みの値 grid に対してシンプルな Excel 数式を評価する。

    サポートする形式:
      =SUM(A1:B3)              範囲の合計
      =SUM(A1:B3, C5:D7)      複数範囲の合計
      =A1                      単一セル参照
      =A1+B1  / =A1-B1  / =A1*B1   2セル間の四則演算
    数式が未サポートの場合、または必要なセルがまだ None の場合は None を返す。
    """
    if not formula or not isinstance(formula, str):
        return None
    expr = formula.lstrip("=").strip()

    def _get(r: int, c: int) -> Optional[float]:
        if 1 <= r < len(grid) and 1 <= c < len(grid[r]):
            v = grid[r][c]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
        return None

    def _range_sum(ref1: str, ref2: str) -> Optional[float]:
        r1, c1 = _parse_cell_ref(ref1)
        r2, c2 = _parse_cell_ref(ref2)
        if None in (r1, c1, r2, c2):
            return None
        total = 0.0
        found = False
        for r in range(min(r1, r2), max(r1, r2) + 1):
            for c in range(min(c1, c2), max(c1, c2) + 1):
                v = _get(r, c)
                if v is not None:
                    total += v
                    found = True
        return total if found else None

    # SUM(arg, arg, ...) — 各引数は範囲参照 ref1:ref2 または単一セル参照
    m = re.fullmatch(r"SUM\((.+)\)", expr, re.IGNORECASE)
    if m:
        args = [a.strip() for a in m.group(1).split(",")]
        grand = 0.0
        found = False
        for arg in args:
            parts = arg.split(":")
            v = _range_sum(parts[0], parts[1]) if len(parts) == 2 else None
            if v is None and len(parts) == 1:
                r, c = _parse_cell_ref(parts[0])
                v = _get(r, c) if r else None
            if v is not None:
                grand += v
                found = True
        return grand if found else None

    # 単一セル参照
    if re.fullmatch(r"\$?[A-Z]+\$?\d+", expr, re.IGNORECASE):
        r, c = _parse_cell_ref(expr)
        return _get(r, c) if r else None

    # 2セル間の四則演算
    m = re.fullmatch(
        r"(\$?[A-Z]+\$?\d+)\s*([+\-\*])\s*(\$?[A-Z]+\$?\d+)", expr, re.IGNORECASE
    )
    if m:
        r1, c1 = _parse_cell_ref(m.group(1))
        op = m.group(2)
        r2, c2 = _parse_cell_ref(m.group(3))
        v1 = _get(r1, c1) if r1 else None
        v2 = _get(r2, c2) if r2 else None
        if v1 is not None and v2 is not None:
            return v1 + v2 if op == "+" else (v1 - v2 if op == "-" else v1 * v2)

    return None


def _fill_formula_cells(
    grid: List[List[Any]], max_row: int, max_col: int, ws_formulas
) -> None:
    """
    *ws_formulas*（data_only=False でロードされたワークシート）の Excel 数式を評価し、
    *grid* 内の None セルを埋める。

    小計が他の小計を合計するケースを正しく解決できるよう、最大5パス実行する
    （各パスで直前まで None だったセルが次のパスの数式で参照可能になる）。
    """
    formula_map: Dict[Tuple[int, int], str] = {}
    for row in ws_formulas.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().startswith("="):
                formula_map[(cell.row, cell.column)] = v.strip()

    for _ in range(5):
        progress = False
        for (r, c), formula in formula_map.items():
            if grid[r][c] is None:
                result = _eval_formula(formula, grid)
                if result is not None:
                    grid[r][c] = result
                    progress = True
        if not progress:
            break  # 新たに解決されたセルがなければ早期終了


# ---------------------------------------------------------------------------
# Grid 構築
# ---------------------------------------------------------------------------


def _build_value_grid(ws, ws_formulas=None) -> Tuple[List[List[Any]], int, int]:
    """
    openpyxl のワークシートから 1-indexed の2次元 grid を構築する。

    - 結合セルの値は結合範囲全体に伝播される。
    - *ws_formulas*（data_only=False でロードした同じシート）が指定された場合、
      キャッシュ値が None のセルを数式文字列から再評価する。
      LibreOffice / Google Sheets からエクスポートされた、数式再計算なしで
      保存されたファイルに対応するための処理。
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        return [[]], 0, 0

    grid: List[List[Any]] = [[None] * (max_col + 1) for _ in range(max_row + 1)]

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                grid[cell.row][cell.column] = cell.value

    # 結合セルの左上の値を結合範囲全体に伝播する
    for merge_range in ws.merged_cells.ranges:
        top_val = grid[merge_range.min_row][merge_range.min_col]
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                grid[r][c] = top_val

    # キャッシュされていない数式セルを評価する（例: キャッシュが欠落している SUM 行）
    if ws_formulas is not None:
        _fill_formula_cells(grid, max_row, max_col, ws_formulas)

    return grid, max_row, max_col


def _build_grid_from_xlrd(sheet) -> Tuple[List[List[Any]], int, int]:
    """xlrd のシートから 1-indexed の2次元 grid を構築する。"""
    import xlrd

    nrows, ncols = sheet.nrows, sheet.ncols
    grid: List[List[Any]] = [[None] * (ncols + 1) for _ in range(nrows + 1)]

    for r in range(nrows):
        for c in range(ncols):
            cell = sheet.cell(r, c)
            if cell.ctype not in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                grid[r + 1][c + 1] = cell.value

    return grid, nrows, ncols


# ---------------------------------------------------------------------------
# テーブル境界の検出
# ---------------------------------------------------------------------------


def _is_filled(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


# 数値パース前に除去する文字（日本語マイナス記号、桁区切り文字など）
_NUM_STRIP = str.maketrans("", "", ",、△▲")

# 日本の統計資料で使われる秘匿・欠損値マーカー。
# 行分類では数値扱いする（テキストとして計上しない）ことで、
# これらが多い行が誤って _RT_COL_HDR に分類されるのを防ぐ。
_STAT_PLACEHOLDERS: frozenset = frozenset({"***", "X", "x", "ｘ", "Ｘ", "－", "…"})



def _cell_is_numeric(v: Any) -> bool:
    """セルの値が数値として解釈できる場合に True を返す。
    統計秘匿マーカー（***、X 等）も数値扱いとする。"""
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return v == v  # float NaN を除外
    s = str(v).strip()
    if s in _STAT_PLACEHOLDERS:
        return True
    try:
        float(s.translate(_NUM_STRIP))
        return True
    except (ValueError, TypeError):
        return False




# ── 行タイプ定数 ────────────────────────────────────────────────────────
_RT_EMPTY   = "empty"    # 入力済みセルなし
_RT_TITLE   = "title"    # 1つの短いテキストセル — テーブル/セクション名の候補
_RT_COL_HDR = "col_hdr"  # テキストセル2つ以上、数値少なめ — 列軸ラベル
_RT_DATA    = "data"     # 55%以上が数値 — 計測値の行
_RT_MIXED   = "mixed"    # 明確に分類できない行（例: 行ヘッダー＋数値）
_RT_NOTE    = "note"     # テーブル境界外の注釈 / 脚注

# セルを注釈 / 脚注として明確にマークするプレフィックス
_NOTE_PREFIXES: Tuple[str, ...] = (
    "※", "＊", "*", "注）", "注)", "（注", "(注", "注意", "＜注", "<注",
    "note:", "NOTE:", "＊注", "*注",
)

# 列挙注釈末尾の集計・関係キーワードにマッチする
# 例: "C-1・C-2・C-3の合計", "A・B・Cの内訳"
_AGG_ENUM_RE = re.compile(
    r"[・、,＋+].+[のをにおける]*(合計|内訳|合算|総計|小計|集計|含む|合わせた|除く|除外)"
)


def _row_content_profile(grid: List[List[Any]], r: int, max_col: int) -> Dict[str, Any]:
    """grid の1行分のコンテンツプロファイル辞書を返す。"""
    numeric = text = 0
    col_min: Optional[int] = None
    col_max: Optional[int] = None
    texts: List[str] = []

    for c in range(1, max_col + 1):
        v = grid[r][c]
        if not _is_filled(v):
            continue
        col_min = c if col_min is None else min(col_min, c)
        col_max = c
        if _cell_is_numeric(v):
            numeric += 1
        else:
            text += 1
            texts.append(str(v).strip())

    return {
        "numeric": numeric,
        "text": text,
        "filled": numeric + text,
        "col_min": col_min,
        "col_max": col_max,
        "texts": texts,
    }


def _classify_row(p: Dict[str, Any]) -> str:
    """行のプロファイルを _RT_* 定数のいずれかに分類する。"""
    filled = p["filled"]
    if filled == 0:
        return _RT_EMPTY

    n_ratio = p["numeric"] / filled

    # 大半が数値 → データ行
    if n_ratio >= 0.55:
        return _RT_DATA

    # テキストセルが1つ — タイトル・注釈/脚注・列ヘッダーを区別する
    if filled == 1 and p["text"] == 1:
        txt = p["texts"][0] if p["texts"] else ""
        # 注釈マーカーまたは非常に長い文 → テーブル外の脚注/注記
        if txt.startswith(_NOTE_PREFIXES) or len(txt) > 60:
            return _RT_NOTE
        """標準の注記プレフィックスなしの列挙＋集計キーワード
        例: "C-1・C-2・C-3の合計", "A区・B区の内訳"
        """
        if _AGG_ENUM_RE.search(txt):
            return _RT_NOTE
        return _RT_TITLE

    """入力済みセルがすべて同一テキスト → 行をまたぐ結合セル。
    Excel では結合セルがマスターセルの値を範囲内のすべての列に伝播する。
    テキストが注釈に見える場合（注記プレフィックス OR 60文字超）は
    _RT_NOTE に分類し、ステートマシンがテーブル領域から除外できるようにする。
    """
    if p["numeric"] == 0 and p["text"] >= 2:
        unique = set(p["texts"])
        if len(unique) == 1:
            txt = next(iter(unique))
            if txt.startswith(_NOTE_PREFIXES) or len(txt) > 60:
                return _RT_NOTE
            if _AGG_ENUM_RE.search(txt):
                return _RT_NOTE
            """4セル以上が同一の短いテキスト → 結合セルによる単位ラベル行
            （例: '百万円'×19, 'mil. yen'×19）。タイトルではなく列ヘッダーとして扱う。
            """
            if p["text"] >= 4:
                return _RT_COL_HDR
            return _RT_TITLE if len(txt) <= 40 else _RT_COL_HDR

    """非常に長いテキストセルと少数の数値を含む行 →
    注釈/注記。別の列の小さな値に隣接した結合セル内の注記を捕捉する
    （「ユニークテキスト」の処理パスが機能しない場合）。
    """
    if p["texts"] and max(len(t) for t in p["texts"]) > 80 and n_ratio <= 0.30:
        return _RT_NOTE

    # 複数のテキストセル、数値が非常に少ない → 列ヘッダー行
    if p["text"] >= 2 and n_ratio <= 0.30:
        return _RT_COL_HDR

    return _RT_MIXED


def _find_column_groups(
    grid: List[List[Any]],
    start_row: int,
    end_row: int,
    max_col: int,
    gap_threshold: int = 2,
) -> List[Tuple[int, int]]:
    """行スパン内で列の空白を境に (start_col, end_col) のペアを返す。

    同じ行範囲を共有する横並びのテーブルを分離するために使用する。
    """
    col_has_content = [False] * (max_col + 1)
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            if _is_filled(grid[r][c]):
                col_has_content[c] = True

    groups: List[Tuple[int, int]] = []
    group_start: Optional[int] = None
    empty_streak = 0

    for c in range(1, max_col + 1):
        if col_has_content[c]:
            if group_start is None:
                group_start = c
            empty_streak = 0
        else:
            if group_start is not None:
                empty_streak += 1
                if empty_streak >= gap_threshold:
                    groups.append((group_start, c - empty_streak))
                    group_start = None
                    empty_streak = 0

    if group_start is not None:
        last_filled = max(
            (c for c in range(group_start, max_col + 1) if col_has_content[c]),
            default=group_start,
        )
        groups.append((group_start, last_filled))

    return groups


def _detect_table_regions(
    grid: List[List[Any]], max_row: int, max_col: int
) -> List[Dict[str, Any]]:
    """構造的な行ごとの分析を使用してテーブル候補領域を検出する。

    アルゴリズム
    ---------
    1. すべての行を empty / title / col_hdr / data / mixed に分類する。
    2. 行を領域に蓄積するステートマシンを実行する。
       単純な「空行 = テーブル境界」を超えた主要な遷移ルール:

       • データ行の後に空の区切りなしで col_hdr 行が現れた場合、
         構造的な不連続性（新しいテーブルヘッダー）を示す
         → 現在の領域をフラッシュして新しい領域を開始する。

       • データ行の後にタイトル行が現れた場合も新しいテーブルを示す。

       • タイトルとそのテーブルのヘッダー/データの間の連続した空行は
         最大4行まで許容する。4行を超えると保留中のタイトルをクリアする。

    3. 各領域内では、値エリア左側の行ヘッダー列が常に含まれるよう、
       すべての構成行から実際の列スパンを追跡する。

    返り値は領域辞書のリスト。各辞書のキー:
        band_start  – 領域の最初の行インデックス
        band_end    – 最後の行インデックス
        col_start   – 全領域行中の最も左の非空列
        col_end     – 最も右の非空列
        title_rows  – [(row_idx, text)] 領域内またはその上のタイトル行
        header_rows – [row_idx] col_hdr 行
        data_rows   – [row_idx] data/mixed 行
    """
    if max_row == 0 or max_col == 0:
        return []

    # ── フェーズ1: すべての行を分類 ──
    profiles: Dict[int, Dict] = {}
    row_types: Dict[int, str] = {}
    for r in range(1, max_row + 1):
        p = _row_content_profile(grid, r, max_col)
        profiles[r] = p
        row_types[r] = _classify_row(p)

    # ── フェーズ2: ステートマシン ──
    def _mk() -> Dict:
        return {
            "band_start": None,
            "band_end": None,
            "col_start": max_col + 1,
            "col_end": 0,
            "title_rows": [],      # [(row_idx, text)]（タイトル行）
            "header_rows": [],     # [row_idx]（ヘッダー行）
            "data_rows": [],       # [row_idx]（データ行）
            "trailing_notes": [],  # このテーブルに後続する注記テキスト
        }

    def _upd_cols(reg: Dict, r: int) -> None:
        p = profiles[r]
        if p["col_min"] is not None:
            reg["col_start"] = min(reg["col_start"], p["col_min"])
        if p["col_max"] is not None:
            reg["col_end"] = max(reg["col_end"], p["col_max"])

    def _flush(reg: Dict, last_row: int, regions: List) -> None:
        if not (reg["data_rows"] or len(reg["header_rows"]) >= 2):
            return
        if reg["band_start"] is None or reg["col_end"] < reg["col_start"]:
            return
        reg["band_end"] = last_row
        regions.append(reg)

    regions: List[Dict] = []
    cur = _mk()
    pending_titles: List[Tuple[int, str]] = []  # まだ領域に紐付けられていないタイトル行
    consec_empty = 0
    last_filled = 0

    for r in range(1, max_row + 1):
        rt = row_types[r]

        # ── 空行 ──────────────────────────────────────────────────────
        if rt == _RT_EMPTY:
            consec_empty += 1
            if cur["data_rows"] and consec_empty >= 2:
                """データ後に連続した空行が2行 → 明確なテーブル境界。
                空行1行は許容する: 複雑なテーブルでは空のサブ行（例: 住宅用 / 学校向け行）
                を持つことがあり、閾値を1にすると誤ってテーブルを分割してしまう。
                注: 空行1行の後に新しいタイトル/幅広ヘッダーが続く場合は、
                _RT_TITLE / _RT_COL_HDR のハンドラーでテーブルを閉じる。
                """
                _flush(cur, last_filled, regions)
                cur = _mk()
                pending_titles = []
            elif consec_empty > 4:
                # 長い空白区間 → 保留中のタイトルを破棄
                pending_titles = []
            continue

        consec_empty = 0
        last_filled = r
        p = profiles[r]

        # ── タイトル行 ──────────────────────────────────────────────────────
        if rt == _RT_TITLE:
            if cur["data_rows"]:
                # データ行の後にタイトルが現れた → 現在のテーブルを終了
                _flush(cur, r - 1, regions)
                cur = _mk()
                pending_titles = []
            elif len(cur["header_rows"]) >= 2:
                """データなしで COL_HDR 行が2行以上あった後にタイトルが来た →
                直前の行が独立したブロックを形成している（例: 実テーブル上の
                ナビゲーション/インデックスセクション）。
                品質フィルタリングで破棄できるようフラッシュする（シグナル3: 短い＋全テキスト＋幅広）。
                """
                _flush(cur, r - 1, regions)
                cur = _mk()
                pending_titles = []
            text = " ".join(p["texts"])
            pending_titles.append((r, text))
            continue

        # ── 列ヘッダー行 ──────────────────────────────────────────────
        if rt == _RT_COL_HDR:
            if cur["data_rows"]:
                """フラッシュ前に、列ヘッダーに見えるだけの小計/集計行でないか確認する。
                テーブルに留める2つのシグナル:
                (a) スパンがテーブル幅の50%未満（狭い小計ラベル）。
                (b) テーブルの列範囲内の充填率が非常に低い — ラベル列が左にあり
                    単一の値が右に離れたデータ行は、スパン的には「幅広」に見えるが
                    実際はスパースなデータ行（例: "住宅用 | … | 1 | … | blank"）。
                """
                tbl_width = cur["col_end"] - cur["col_start"] + 1
                row_width = (
                    p["col_max"] - p["col_min"] + 1
                    if p["col_min"] is not None and p["col_max"] is not None
                    else 0
                )
                filled_in_span = (
                    sum(
                        1 for c in range(cur["col_start"], cur["col_end"] + 1)
                        if _is_filled(grid[r][c])
                    )
                    if tbl_width > 0
                    else 0
                )
                fill_density = filled_in_span / tbl_width if tbl_width > 0 else 0

                if tbl_width > 0 and (
                    row_width / tbl_width < 0.5 or fill_density < 0.30
                ):
                    cur["data_rows"].append(r)
                    _upd_cols(cur, r)
                    continue

                """データの後に空行なしでヘッダーが現れた
                → 構造的な不連続: 現在を閉じ、新しく開始する
                """
                _flush(cur, r - 1, regions)
                cur = _mk()
                # pending_titles はこの新しいテーブルに属する可能性があるため保持する

            if cur["band_start"] is None:
                # Attach pending titles and open the region
                for tr, tt in pending_titles:
                    cur["title_rows"].append((tr, tt))
                    _upd_cols(cur, tr)
                pending_titles = []
                cur["band_start"] = cur["title_rows"][0][0] if cur["title_rows"] else r

            cur["header_rows"].append(r)
            _upd_cols(cur, r)
            continue

        # ── 注釈 / 脚注行 ─────────────────────────────────────
        if rt == _RT_NOTE:
            # 重複除去: 結合セルは同じテキストをすべての列に伝播する
            unique_note_texts = list(dict.fromkeys(profiles[r]["texts"]))
            note_text = unique_note_texts[0] if unique_note_texts else ""
            if note_text:
                if cur["data_rows"]:
                    # 注記が現在のテーブルに後続する → まずテーブルをフラッシュし、
                    # 領域境界が注記の前に終わるよう注記を付加する。
                    _flush(cur, r - 1, regions)
                    if regions:
                        regions[-1]["trailing_notes"].append(note_text)
                    cur = _mk()
                    pending_titles = []
                elif cur["band_start"] is None and regions:
                    # 新しいテーブルはまだ開始していない — 直前にフラッシュされた
                    # テーブルに後続する連続した注記行。
                    regions[-1]["trailing_notes"].append(note_text)
                # else: テーブルの前の注記（前の領域なし）— スキップ
            continue

        # ── データ / mixed 行 ───────────────────────────────────────────────
        if cur["band_start"] is None:
            for tr, tt in pending_titles:
                cur["title_rows"].append((tr, tt))
                _upd_cols(cur, tr)
            pending_titles = []
            cur["band_start"] = cur["title_rows"][0][0] if cur["title_rows"] else r

        cur["data_rows"].append(r)
        _upd_cols(cur, r)

    # 最終フラッシュ
    if cur["band_start"] is not None:
        _flush(cur, last_filled, regions)

    return regions


def _classify_table_quality(
    df: "pd.DataFrame",
    grid: List[List[Any]],
    band_start: int,
    band_end: int,
    col_start: int,
    col_end: int,
) -> str:
    """検出された矩形領域を3つの品質ティアのいずれかに分類する。

    Returns
    -------
    "ok"       — 分析に適している
    "metadata" — 構造化されているが分析的でない（メニュー、インデックス/目次テーブル）
    "discard"  — テーブルではない（フリーテキスト段落、注記、空のフラグメント）

    シグナルマッピング
    --------------
    1. 充填率 < 20%                            → discard
    2. 密度の高いデータ行が2行未満             → discard
    3. 幅広＋短い＋全テキスト（≤5行, ≥5列, 数値0） → discard
    4. 長いテキスト（列ヘッダーまたはセルが80文字超、または25%超のテキストセルが40文字超）
       • ＋数値セルがゼロ                      → discard（フリーテキストブロック）
       • ＋数値セルあり                        → discard（インデックス/目次テーブル）
    """
    # シグナル1: 生の充填率
    raw_total = (band_end - band_start + 1) * (col_end - col_start + 1)
    if raw_total == 0:
        return "discard"
    raw_filled = sum(
        1
        for r in range(band_start, band_end + 1)
        for c in range(col_start, col_end + 1)
        if _is_filled(grid[r][c])
    )
    if raw_filled / raw_total < 0.20:
        return "discard"

    # シグナル2: 密度の高い行（列の25%以上が充填済み）
    if df.empty:
        return "discard"
    n_cols = max(len(df.columns), 1)
    dense_rows = sum(
        1
        for _, row in df.iterrows()
        if sum(1 for v in row if v is not None and str(v).strip()) / n_cols >= 0.25
    )
    if dense_rows < 2:
        return "discard"

    # シグナル3: 幅広＋短い＋全テキスト → 選択メニュー / フォームグリッド
    if df.shape[0] <= 5 and df.shape[1] >= 5:
        num_ct = sum(
            int(pd.to_numeric(df[c], errors="coerce").notna().sum()) for c in df.columns
        )
        if num_ct == 0:
            return "discard"

    # シグナル4: 長いテキストコンテンツ（列ヘッダーとセル値の両方をチェック）
    txt_ct = lng_ct = 0
    max_len = 0

    def _tally(raw: str) -> None:
        nonlocal txt_ct, lng_ct, max_len
        s = raw.strip()
        if not s or s.lower() == "nan":
            return
        try:
            float(s.translate(_NUM_STRIP))
            return
        except (ValueError, TypeError):
            pass
        txt_ct += 1
        n = len(s)
        if n > max_len:
            max_len = n
        if n > 40:
            lng_ct += 1

    for col_name in df.columns:
        _tally(str(col_name))
    for _, row in df.iterrows():
        for v in row:
            if not pd.isna(v):
                _tally(str(v))

    if max_len > 80 or (txt_ct >= 3 and lng_ct / txt_ct > 0.25):
        """ガード: 迷い込んだ注釈行が領域境界に入り込んだだけで
        データの豊富なテーブルを破棄してはならない。
        絶対的な数値カウントが十分に大きい場合、または数値密度
        （数値セルの割合）が有意な場合にテーブルを保持する —
        密度チェックは絶対カウント < 10 の小さなテーブルを救済する。
        """
        total_numeric = sum(
            int(pd.to_numeric(df[c], errors="coerce").notna().sum())
            for c in df.columns
        )
        numeric_density = total_numeric / raw_total if raw_total > 0 else 0
        if total_numeric < 10 and numeric_density < 0.15:
            return "discard"

    return "ok"


# ---------------------------------------------------------------------------
# grid 領域からの DataFrame 抽出
# ---------------------------------------------------------------------------




def _make_unique_columns(columns: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for col in columns:
        col = col if col else "列"
        if col in seen:
            seen[col] += 1
            result.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            result.append(col)
    return result




def _extract_dataframe(
    grid: List[List[Any]],
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> Tuple[pd.DataFrame, Optional[str], Optional[pd.DataFrame]]:
    """矩形領域を DataFrame として抽出する。

    Returns (df, title, raw_df):
      df      — 整形済み DataFrame（多段ヘッダーをマージ済み）
      title   — テーブル直上のセクションタイトルテキスト
      raw_df  — 整形前 DataFrame（多段ヘッダー時のみ設定、それ以外は None）
    """
    if start_row > end_row or start_col > end_col:
        return pd.DataFrame(), None, None

    rows = [
        [grid[r][c] for c in range(start_col, end_col + 1)]
        for r in range(start_row, end_row + 1)
    ]

    if not rows:
        return pd.DataFrame(), None, None

    n_title, header_roles = detect_header_roles(rows)
    num_cols = end_col - start_col + 1

    # スキップされたタイトル行からタイトルテキストを収集する
    title: Optional[str] = None
    if n_title > 0:
        title_parts = []
        for i in range(n_title):
            vals = [v for v in rows[i] if v is not None]
            if vals:
                title_parts.append(str(vals[0]).strip())
        title = " / ".join(title_parts) if title_parts else None

    remaining = rows[n_title:]
    n_header = len(header_roles)

    if n_header == 0:
        columns = [f"列{i + 1}" for i in range(num_cols)]
        df = pd.DataFrame(remaining, columns=columns)
        return df.dropna(how="all").reset_index(drop=True), title, None

    if n_header == 1:
        header = _make_unique_columns(
            [str(v) if v is not None else "" for v in remaining[0]]
        )
        df = pd.DataFrame(remaining[1:], columns=header)
        return df.dropna(how="all").reset_index(drop=True), title, None

    # 多段ヘッダー: 整形前 DataFrame を raw_df として保存
    raw_header = _make_unique_columns(
        [str(v) if v is not None else "" for v in remaining[0]]
    )
    raw_df = pd.DataFrame(remaining[1:], columns=raw_header).dropna(how="all").reset_index(drop=True)

    header_data = [remaining[i] for i in range(n_header)]
    merged = merge_header_rows(header_data, header_roles, num_cols)
    header = _make_unique_columns(merged)
    df = pd.DataFrame(remaining[n_header:], columns=header)

    return df.dropna(how="all").reset_index(drop=True), title, raw_df


# ---------------------------------------------------------------------------
# ワークシートごとの検出
# ---------------------------------------------------------------------------


def _detect_tables_in_grid(
    grid: List[List[Any]],
    max_row: int,
    max_col: int,
    sheet_name: str,
    table_counter: Dict[str, int],
) -> List[DetectedTable]:
    """構造的な行分類とステートマシンを使用してテーブルを検出する。"""
    if max_row == 0 or max_col == 0:
        return []

    regions = _detect_table_regions(grid, max_row, max_col)
    detected: List[DetectedTable] = []

    for reg in regions:
        band_start = reg["band_start"]
        band_end = reg["band_end"]

        col_groups = _find_column_groups(
            grid, band_start, band_end, max_col, gap_threshold=2
        )

        """タイトル行をスキップし、最初の列ヘッダー行をテーブル開始行とする。
        これにより start_row がシート内の実データヘッダー位置（例: 行8や行10）に
        一致し、タイトル行（行1〜7等）は title フィールドへ格納される。
        """
        first_header_row = min(reg["header_rows"]) if reg["header_rows"] else band_start

        for col_start, col_end in col_groups:
            if (band_end - first_header_row + 1) < 2:
                continue

            df, inner_title, raw_df = _extract_dataframe(
                grid, first_header_row, band_end, col_start, col_end
            )
            if df.empty or len(df) == 0:
                continue

            quality = _classify_table_quality(
                df, grid, first_header_row, band_end, col_start, col_end
            )
            if quality != "ok":
                continue

            effective_title = inner_title
            if effective_title is None and reg["title_rows"]:
                effective_title = " / ".join(t for _, t in reg["title_rows"])

            safe = sheet_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
            table_counter[safe] = table_counter.get(safe, 0) + 1
            table_id = f"{safe}_T{table_counter[safe]}"

            # グルーピング列の前方補完（視覚結合セル対応）
            pre_fill_df_candidate = df
            df, filled_cols = fill_grouping_cols(df)
            pre_fill_df = pre_fill_df_candidate if filled_cols else None

            # 集計行・列を除去
            cleaned_df, agg_rows, agg_cols, agg_row_positions = remove_aggregates(df)
            pre_agg_df = df if (agg_rows or agg_cols) else None

            detected.append(
                DetectedTable(
                    table_id=table_id,
                    sheet_name=sheet_name,
                    start_row=first_header_row,
                    end_row=band_end,
                    start_col=col_start,
                    end_col=col_end,
                    df=cleaned_df,
                    title=effective_title,
                    notes=reg.get("trailing_notes", []),
                    raw_df=raw_df,
                    pre_agg_df=pre_agg_df,
                    agg_rows_removed=agg_rows,
                    agg_cols_removed=agg_cols,
                    agg_rows_removed_positions=agg_row_positions,
                    filled_cols=filled_cols,
                    pre_fill_df=pre_fill_df,
                )
            )

    _propagate_sheet_title(detected)
    return detected


_YEAR_RE = re.compile(r"\d{4}")
# タイトルが年次のみ（例: "2012年", "2012年度", "2012"）かを判定する正規表現
_SECTION_ONLY_RE = re.compile(r"^\s*\d{4}[年度]?\s*$")


def _propagate_sheet_title(detected: List[DetectedTable]) -> None:
    """
    同一シート内で複数テーブルを検出した際、先頭テーブルタイトルの
    シート共通部分（年次パターンを含まない要素）を後続テーブルに引き継ぐ。

    安全条件（両方を満たす場合のみ動作）:
      1. 先頭テーブルのタイトルに " / " が含まれる
         → 階層構造（シート共通タイトル / セクション）が確認できる場合のみ
      2. 後続テーブルのタイトルが年次のみ（\d{4}年? 形式）
         → 独自のサブタイトルを持つテーブルは除外

    例:
      T1: "シート共通タイトル / 2011年" → sheet_title = "シート共通タイトル"
      T2: "2012年" → "シート共通タイトル_2012年"  ← 引き継ぎ対象
      T2: "サブタイトル / 2012年" → そのまま      ← 引き継ぎ対象外
    """
    if not detected:
        return

    from itertools import groupby
    for _sheet, group_iter in groupby(detected, key=lambda t: t.sheet_name):
        tables = list(group_iter)
        if len(tables) < 2:
            continue

        first_title = tables[0].title or ""

        # 条件1: 先頭タイトルに " / " が含まれる（階層構造が確認できる）
        if " / " not in first_title:
            continue

        parts = [p.strip() for p in first_title.split("/")]
        sheet_parts = [p for p in parts if not _YEAR_RE.search(p) and p]
        section_parts = [p for p in parts if _YEAR_RE.search(p) and p]

        if not sheet_parts:
            continue

        sheet_title = " ".join(sheet_parts)

        # 後続テーブルを先に確認し、引き継ぎが発生するかを判定
        propagated = False
        for t in tables[1:]:
            raw = t.title or ""
            # 条件2: タイトルが年次だけ（4桁数字 + オプションで年/度）
            if not _SECTION_ONLY_RE.match(raw):
                continue  # 独立したタイトルを持つ → 引き継ぎしない
            t.title = f"{sheet_title}_{raw.strip()}"
            propagated = True

        # 先頭テーブルも "_" 区切りに整形（引き継ぎが実際に発生した場合のみ）
        if propagated and section_parts:
            tables[0].title = f"{sheet_title}_{section_parts[0]}"


def _apply_cross_table_detection(tables: List[DetectedTable], filename: str) -> None:
    """テーブルリスト全体にクロス集計検出と縦持ち変換を適用する。"""
    for t in tables:
        if t.df is None or t.df.empty:
            continue
        info = detect_cross_table(t.df, title=t.title, filename=filename)
        if info:
            t.stack_info = info
            t.stacked_df = stack_cross_table(t.df, info)


# ---------------------------------------------------------------------------
# 公開 API
# ---------------------------------------------------------------------------


def parse_excel(
    file_content: bytes, filename: str
) -> Tuple[List[DetectedTable], List[str]]:
    """.xlsx/.xlsm ファイルを解析し、すべてのテーブル領域を検出する。"""
    ext = Path(filename).suffix.lower()
    all_tables: List[DetectedTable] = []
    table_counter: Dict[str, int] = {}

    if ext in (".xlsx", ".xlsm", ".xlsb"):
        raw = io.BytesIO(file_content)
        wb_data = openpyxl.load_workbook(raw, data_only=True, read_only=False)
        raw.seek(0)
        wb_form = openpyxl.load_workbook(raw, data_only=False, read_only=False)
        sheet_names = wb_data.sheetnames

        for sheet_name in sheet_names:
            ws_data = wb_data[sheet_name]
            ws_form = wb_form[sheet_name] if sheet_name in wb_form.sheetnames else None
            grid, max_row, max_col = _build_value_grid(ws_data, ws_formulas=ws_form)
            tables = _detect_tables_in_grid(
                grid, max_row, max_col, sheet_name, table_counter
            )
            all_tables.extend(tables)

    elif ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(file_contents=file_content)
        sheet_names = wb.sheet_names()

        for sheet_name in sheet_names:
            ws = wb.sheet_by_name(sheet_name)
            grid, max_row, max_col = _build_grid_from_xlrd(ws)
            tables = _detect_tables_in_grid(
                grid, max_row, max_col, sheet_name, table_counter
            )
            all_tables.extend(tables)
    else:
        raise ValueError(f"Unsupported Excel format: {ext}")

    _apply_cross_table_detection(all_tables, filename)
    return all_tables, sheet_names


def parse_csv(
    file_content: bytes, filename: str
) -> Tuple[List[DetectedTable], List[str]]:
    """CSV ファイルを単一のテーブルとして解析する。"""
    for enc in ("utf-8-sig", "cp932", "shift_jis", "utf-8"):
        try:
            df = pd.read_csv(io.BytesIO(file_content), encoding=enc)
            break
        except (UnicodeDecodeError, Exception):
            continue
    else:
        df = pd.read_csv(io.BytesIO(file_content), encoding="latin-1")

    safe = Path(filename).stem.replace(" ", "_")
    table = DetectedTable(
        table_id=f"{safe}_T1",
        sheet_name="CSV",
        start_row=1,
        end_row=len(df) + 1,
        start_col=1,
        end_col=len(df.columns),
        df=df,
    )
    _apply_cross_table_detection([table], filename)
    return [table], ["CSV"]
