"""
ステップ1 ファイル読み込みモジュール。

処理概要: ファイルのバイト列を受け取り、セル値を保持した生グリッド（1-indexed 2次元配列）を構築する。
          数式評価・結合セル伝播を行う。CSV も同じ grid 形式で返すことで、
          Step2 のテーブル検出（タイトル行・ヘッダー行の認識）が Excel / CSV
          共通の grid ベースパイプラインで処理できるようにする。
          テーブル検出・整形・分析は行わない。
入力    : bytes（Excel / CSV ファイルの内容）、ファイル名
出力    : List[SheetGrid]（シートごとの生グリッド）
"""

import io
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from .models import SheetGrid


# ---------------------------------------------------------------------------
# 数式評価ヘルパー
# openpyxl を data_only=True でロードすると、Excel で一度も開かれていないファイルの
# 数式セルは None になる。そのままでは step2 にスカスカのグリッドが渡るため、
# ここで数式を自前評価してセル値を補完する（「ファイルの値を正確に取り出す」読み込みの一部）。
# ---------------------------------------------------------------------------

def _col_to_num(col: str) -> int:
    """'A'→1, 'Z'→26, 'AA'→27（大文字・小文字を区別しない）。"""
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _parse_cell_ref(ref: str) -> Tuple[Optional[int], Optional[int]]:
    """'D5', '$D$5' → (row=5, col=4)。失敗時は (None, None) を返す。"""
    m = re.match(r"\$?([A-Z]+)\$?(\d+)$", ref.strip().upper())
    if m:
        return int(m.group(2)), _col_to_num(m.group(1))
    return None, None


def _eval_formula(formula: str, grid: List[List[Any]]) -> Optional[float]:
    """
    構築済みの値 grid に対してシンプルな Excel 数式を評価する。

    サポートする形式:
      =SUM(A1:B3)              範囲の合計
      =SUM(A1:B3, C5:D7)      複数範囲の合計
      =A1                      単一セル参照
      =A1+B1  / =A1-B1  / =A1*B1   2セル間の四則演算
    数式が未サポートの場合、または必要なセルがまだ None の場合は None を返す。
    """
    if not formula or not isinstance(formula, str):
        return None
    expr = formula.lstrip("=").strip()

    def _get(r: int, c: int) -> Optional[float]:
        if 1 <= r < len(grid) and 1 <= c < len(grid[r]):
            v = grid[r][c]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
        return None

    def _range_sum(ref1: str, ref2: str) -> Optional[float]:
        r1, c1 = _parse_cell_ref(ref1)
        r2, c2 = _parse_cell_ref(ref2)
        if None in (r1, c1, r2, c2):
            return None
        total = 0.0
        found = False
        for r in range(min(r1, r2), max(r1, r2) + 1):
            for c in range(min(c1, c2), max(c1, c2) + 1):
                v = _get(r, c)
                if v is not None:
                    total += v
                    found = True
        return total if found else None

    m = re.fullmatch(r"SUM\((.+)\)", expr, re.IGNORECASE)
    if m:
        args = [a.strip() for a in m.group(1).split(",")]
        grand = 0.0
        found = False
        for arg in args:
            parts = arg.split(":")
            v = _range_sum(parts[0], parts[1]) if len(parts) == 2 else None
            if v is None and len(parts) == 1:
                r, c = _parse_cell_ref(parts[0])
                v = _get(r, c) if r else None
            if v is not None:
                grand += v
                found = True
        return grand if found else None

    if re.fullmatch(r"\$?[A-Z]+\$?\d+", expr, re.IGNORECASE):
        r, c = _parse_cell_ref(expr)
        return _get(r, c) if r else None

    m = re.fullmatch(
        r"(\$?[A-Z]+\$?\d+)\s*([+\-\*])\s*(\$?[A-Z]+\$?\d+)", expr, re.IGNORECASE
    )
    if m:
        r1, c1 = _parse_cell_ref(m.group(1))
        op = m.group(2)
        r2, c2 = _parse_cell_ref(m.group(3))
        v1 = _get(r1, c1) if r1 else None
        v2 = _get(r2, c2) if r2 else None
        if v1 is not None and v2 is not None:
            return v1 + v2 if op == "+" else (v1 - v2 if op == "-" else v1 * v2)

    return None


def _fill_formula_cells(
    grid: List[List[Any]], max_row: int, max_col: int, ws_formulas
) -> None:
    """
    *ws_formulas*（data_only=False でロードされたワークシート）の Excel 数式を評価し、
    *grid* 内の None セルを埋める。

    小計が他の小計を合計するケースを正しく解決できるよう、最大5パス実行する。
    """
    formula_map: Dict[Tuple[int, int], str] = {}
    for row in ws_formulas.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().startswith("="):
                formula_map[(cell.row, cell.column)] = v.strip()

    for _ in range(5):
        progress = False
        for (r, c), formula in formula_map.items():
            if grid[r][c] is None:
                result = _eval_formula(formula, grid)
                if result is not None:
                    grid[r][c] = result
                    progress = True
        if not progress:
            break


# ---------------------------------------------------------------------------
# グリッド構築
# ---------------------------------------------------------------------------

def _build_value_grid(ws, ws_formulas=None) -> Tuple[List[List[Any]], int, int]:
    """
    openpyxl のワークシートから 1-indexed の2次元 grid を構築する。

    - 結合セルの値は結合範囲全体に伝播される。
    - *ws_formulas* が指定された場合、キャッシュ値が None のセルを数式から再評価する。
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        return [[]], 0, 0

    grid: List[List[Any]] = [[None] * (max_col + 1) for _ in range(max_row + 1)]

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                grid[cell.row][cell.column] = cell.value

    # openpyxl の内部表現では結合セルの値は先頭セルにしか存在しない。
    # 伝播させないと step2 が結合範囲を空セルと誤判定するため、ここで補正する。
    for merge_range in ws.merged_cells.ranges:
        top_val = grid[merge_range.min_row][merge_range.min_col]
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                grid[r][c] = top_val

    if ws_formulas is not None:
        _fill_formula_cells(grid, max_row, max_col, ws_formulas)

    return grid, max_row, max_col


def _build_grid_from_xlrd(sheet) -> Tuple[List[List[Any]], int, int]:
    """xlrd のシートから 1-indexed の2次元 grid を構築する。"""
    import xlrd

    nrows, ncols = sheet.nrows, sheet.ncols
    grid: List[List[Any]] = [[None] * (ncols + 1) for _ in range(nrows + 1)]

    for r in range(nrows):
        for c in range(ncols):
            cell = sheet.cell(r, c)
            if cell.ctype not in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                grid[r + 1][c + 1] = cell.value

    return grid, nrows, ncols


# ---------------------------------------------------------------------------
# 公開 API
# ---------------------------------------------------------------------------

def load_excel(content: bytes, filename: str) -> Tuple[List[SheetGrid], List[str]]:
    """Excel ファイルを読み込み、シートごとの SheetGrid リストと sheet_names を返す。"""
    ext = Path(filename).suffix.lower()
    sheets: List[SheetGrid] = []

    if ext in (".xlsx", ".xlsm", ".xlsb"):
        raw = io.BytesIO(content)
        wb_data = openpyxl.load_workbook(raw, data_only=True, read_only=False)
        raw.seek(0)
        wb_form = openpyxl.load_workbook(raw, data_only=False, read_only=False)
        sheet_names = wb_data.sheetnames

        for sheet_name in sheet_names:
            ws_data = wb_data[sheet_name]
            ws_form = wb_form[sheet_name] if sheet_name in wb_form.sheetnames else None
            grid, max_row, max_col = _build_value_grid(ws_data, ws_formulas=ws_form)
            sheets.append(SheetGrid(sheet_name=sheet_name, grid=grid, max_row=max_row, max_col=max_col))

    elif ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(file_contents=content)
        sheet_names = wb.sheet_names()

        for sheet_name in sheet_names:
            ws = wb.sheet_by_name(sheet_name)
            grid, max_row, max_col = _build_grid_from_xlrd(ws)
            sheets.append(SheetGrid(sheet_name=sheet_name, grid=grid, max_row=max_row, max_col=max_col))

    else:
        raise ValueError(f"Unsupported Excel format: {ext}")

    return sheets, sheet_names


def _parse_csv_cell(raw: str) -> Any:
    """CSV セル文字列を Excel セルの型付き値に近づける（空文字→None、数値文字列→int/float）。

    CSV は全セルが文字列として読み込まれるため、そのまま grid に積むと Step2 の
    行分類（タイトル行・ヘッダー行・データ行の判定）が Excel のネイティブ型付き
    セル前提の一部ロジックと食い違う。ここで型を復元しておくことで、CSV も
    Excel と全く同じ grid ベースの検出パイプライン（detect_tables）に載せられる。
    """
    s = raw.strip()
    if s == "":
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def _build_grid_from_csv(
    content: bytes,
) -> Tuple[List[List[Any]], int, int, List[int]]:
    """CSV バイト列から 1-indexed の2次元 grid を構築する（Excel の
    _build_value_grid と同じ形の出力。行数・列数が不揃いな行は右側を None 埋めする）。

    行ごとの実際のフィールド数（パディング前、末尾の明示的な空フィールドは
    含む）も row_widths として返す。1つの CSV ファイル内に幅の異なる複数の
    テーブルが含まれる場合、grid 自体はファイル全体で共通の max_col に
    揃えて右側を None 埋めするため、grid の値だけでは「その行の表が本来
    どこまでの幅を持つか」と「たまたま他の行のせいで grid 上は列が
    存在するが、この行にとっては無関係な padding か」を区別できない。
    row_widths は Step2 側でこの区別に使う（末尾の空列をどこまで
    そのテーブル自身の列として扱ってよいかの判定）。
    """
    import csv as _csv

    text: Optional[str] = None
    for enc in ("utf-8-sig", "cp932", "shift_jis", "utf-8"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = content.decode("latin-1")

    raw_rows = list(_csv.reader(io.StringIO(text)))
    rows = [[_parse_csv_cell(cell) for cell in row] for row in raw_rows]
    # 末尾の完全空行を除去する（pd.read_csv の挙動に揃える）
    while rows and all(v is None for v in rows[-1]):
        rows.pop()

    if not rows:
        return [[]], 0, 0, []

    max_row = len(rows)
    max_col = max(len(r) for r in rows)
    grid: List[List[Any]] = [[None] * (max_col + 1) for _ in range(max_row + 1)]
    row_widths: List[int] = [0] * (max_row + 1)
    for i, row in enumerate(rows, start=1):
        row_widths[i] = len(row)
        for j, v in enumerate(row, start=1):
            grid[i][j] = v

    return grid, max_row, max_col, row_widths


def load_csv(content: bytes) -> Tuple[List[SheetGrid], List[str]]:
    """CSV ファイルを読み込み、SheetGrid リストと sheet_names を返す（load_excel と同形）。

    Step2 のテーブル検出（タイトル行・ヘッダー行の認識）が Excel と共通の
    grid ベースパイプライン（detect_tables）で処理できるようにするため、
    pandas の header=0 決め打ちでは表現できない grid 形式で返す。"""
    grid, max_row, max_col, row_widths = _build_grid_from_csv(content)
    sheet = SheetGrid(
        sheet_name="CSV",
        grid=grid,
        max_row=max_row,
        max_col=max_col,
        row_widths=row_widths,
    )
    return [sheet], ["CSV"]
